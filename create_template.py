#!/usr/bin/env python3
"""
Generate the Excel template for a PRISM HIV study dashboard.
Run: python create_template.py
Output: HIV_Study_Template.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SEGMENTS = [
    "TSP","CEC","TC","HF","PP","WE","PFF","HHN","MFL","VS",
    "UCP","FJP","HCP","HAD","HCI","GHI"
]
SEG_NAMES = [
    "Trust the Science Pragmatists","Consumer Empowerment Champions","Traditional Conservatives",
    "Health Futurists","Price Populists","Wellness Evangelists","Paleo Freedom Fighters",
    "Holistic Health Naturalists","Medical Freedom Libertarians","Vaccine Skeptics",
    "Universal Care Progressives","Faith & Justice Progressives","Health Care Protectionists",
    "Health Abundance Democrats","Health Care Incrementalists","Global Health Institutionalists"
]

HEADER_FILL = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="e2e8f0")
INST_FONT = Font(name="Calibri", size=9, italic=True, color="94a3b8")
THIN_BORDER = Border(
    left=Side(style='thin', color='334155'),
    right=Side(style='thin', color='334155'),
    top=Side(style='thin', color='334155'),
    bottom=Side(style='thin', color='334155')
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def add_seg_headers(ws, row, start_col=3):
    ws.cell(row=row, column=1, value="Message ID")
    ws.cell(row=row, column=2, value="Token")
    for i, seg in enumerate(SEGMENTS):
        ws.cell(row=row, column=start_col + i, value=seg)
    style_header(ws, row, start_col + 15)

def add_sop_matrix(ws, start_row, n_messages, n_tokens_per_msg):
    row = start_row
    for msg_id in range(1, n_messages + 1):
        # Aggregate row
        ws.cell(row=row, column=1, value=msg_id)
        ws.cell(row=row, column=2, value="TOTAL")
        row += 1
        # Token rows
        for t in range(n_tokens_per_msg):
            ws.cell(row=row, column=1, value=msg_id)
            ws.cell(row=row, column=2, value=f"Token {t}" if t > 0 else "Core")
            row += 1
    return row

wb = openpyxl.Workbook()

# ─── Tab 1: StudyMeta ───
ws = wb.active
ws.title = "StudyMeta"
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 60
meta_fields = [
    ("study_name", "HIV"),
    ("client", "Gilead"),
    ("topic", "HIV Treatment & Prevention"),
    ("field_date", "2025"),
    ("n_messages", "15"),
    ("n_tokens_per_message", "3"),
    ("methodology", "MaxDiff · 16 PRISM segments"),
]
ws.cell(row=1, column=1, value="Field").font = HEADER_FONT
ws.cell(row=1, column=2, value="Value").font = HEADER_FONT
for i, (k, v) in enumerate(meta_fields, start=2):
    ws.cell(row=i, column=1, value=k)
    ws.cell(row=i, column=2, value=v)

# ─── Tab 2: Messages ───
ws = wb.create_sheet("Messages")
headers = ["msg_id", "token_id", "short_name", "theme", "message_text"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 80
ws.cell(row=2, column=1).font = INST_FONT
ws.cell(row=2, column=1, value="# msg_id: message number (1-N)")
ws.cell(row=3, column=1, value="# token_id: 0=Core, 1=Token1, 2=Token2").font = INST_FONT
ws.cell(row=4, column=1, value="# For aggregate/total row, use token_id=-1").font = INST_FONT

# ─── Tab 3: VariantText ───
ws = wb.create_sheet("VariantText")
headers = ["msg_id", "token_id"] + SEGMENTS
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))
ws.cell(row=2, column=1, value="# One row per message×token combo. Each segment column has that segment's persona variant text.").font = INST_FONT

# ─── Tab 4: ControlSoP ───
ws = wb.create_sheet("ControlSoP")
ws.cell(row=1, column=1, value="# Control condition SoP values. TOTAL row = aggregate across tokens. Then one row per token.").font = INST_FONT
add_seg_headers(ws, 2)

# ─── Tab 5: VariantSoP ───
ws = wb.create_sheet("VariantSoP")
ws.cell(row=1, column=1, value="# Persona variant condition SoP values. Same structure as ControlSoP.").font = INST_FONT
add_seg_headers(ws, 2)

# ─── Tab 6: SigFlags_Control ───
ws = wb.create_sheet("SigFlags_Control")
ws.cell(row=1, column=1, value="# Significance flags for Control tokens vs Core. +1=sig higher, -1=sig lower, 0=not sig. Leave TOTAL/Core rows as 0.").font = INST_FONT
add_seg_headers(ws, 2)

# ─── Tab 7: SigFlags_Variant ───
ws = wb.create_sheet("SigFlags_Variant")
ws.cell(row=1, column=1, value="# Significance flags for Variant tokens vs Core. Same structure.").font = INST_FONT
add_seg_headers(ws, 2)

# ─── Tab 8: SegmentMetrics ───
ws = wb.create_sheet("SegmentMetrics")
headers = ["code", "name", "party", "pop", "roi", "highRoi", "supporters", "activation", "influence",
           "persuad_strong_support", "persuad_lean_support", "persuad_persuadable", "persuad_lean_oppose", "persuad_strong_oppose",
           "prepost_key1_pre", "prepost_key1_post", "prepost_key1_label", "prepost_key1_question", "prepost_key1_scale",
           "prepost_key2_pre", "prepost_key2_post", "prepost_key2_label", "prepost_key2_question", "prepost_key2_scale",
           "prepost_key3_pre", "prepost_key3_post", "prepost_key3_label", "prepost_key3_question", "prepost_key3_scale",
           "prepost_key4_pre", "prepost_key4_post", "prepost_key4_label", "prepost_key4_question", "prepost_key4_scale"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))
for i, (code, name) in enumerate(zip(SEGMENTS, SEG_NAMES), start=2):
    ws.cell(row=i, column=1, value=code)
    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value="GOP" if i <= 11 else "DEM")

# ─── Tab 9: ThemeColors ───
ws = wb.create_sheet("ThemeColors")
ws.cell(row=1, column=1, value="theme").font = HEADER_FONT
ws.cell(row=1, column=2, value="hex_color").font = HEADER_FONT
ws.cell(row=2, column=1, value="# Add theme names and hex colors, e.g. Epidemic=#f87171").font = INST_FONT

fname = "HIV_Study_Template.xlsx"
wb.save(fname)
print(f"Created {fname}")
